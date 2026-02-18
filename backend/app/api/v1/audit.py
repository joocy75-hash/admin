"""Audit log endpoints: list, detail, Excel export."""

import io
from datetime import datetime, date

from fastapi import APIRouter, Depends, Query
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.responses import StreamingResponse

from app.database import get_session
from app.api.deps import PermissionChecker
from app.models.admin_user import AdminUser
from app.models.audit_log import AuditLog
from app.schemas.audit import AuditLogListResponse, AuditLogResponse

router = APIRouter(prefix="/audit", tags=["audit"])

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _build_log_response(session: AsyncSession, log: AuditLog) -> AuditLogResponse:
    admin_user = await session.get(AdminUser, log.admin_user_id) if log.admin_user_id else None
    return AuditLogResponse(
        id=log.id,
        admin_user_id=log.admin_user_id,
        admin_username=admin_user.username if admin_user else None,
        ip_address=log.ip_address,
        user_agent=log.user_agent,
        action=log.action,
        module=log.module,
        resource_type=log.resource_type,
        resource_id=log.resource_id,
        before_data=log.before_data,
        after_data=log.after_data,
        description=log.description,
        created_at=log.created_at,
    )


def _parse_dates(start_date: str | None, end_date: str | None) -> tuple[datetime | None, datetime | None]:
    start = None
    end = None
    if start_date:
        start = datetime.combine(
            datetime.strptime(start_date, "%Y-%m-%d").date(),
            datetime.min.time(),
        )
    if end_date:
        end = datetime.combine(
            datetime.strptime(end_date, "%Y-%m-%d").date(),
            datetime.max.time(),
        )
    return start, end


# ─── List Audit Logs ──────────────────────────────────────────────

@router.get("/logs", response_model=AuditLogListResponse)
async def list_audit_logs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    action: str | None = Query(None),
    module: str | None = Query(None),
    admin_user_id: int | None = Query(None),
    admin_username: str | None = Query(None, description="Admin username (partial match)"),
    start_date: str | None = Query(None, description="YYYY-MM-DD"),
    end_date: str | None = Query(None, description="YYYY-MM-DD"),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("audit_log.view")),
):
    base = select(AuditLog)

    if action:
        base = base.where(AuditLog.action == action)
    if module:
        base = base.where(AuditLog.module == module)
    if admin_user_id:
        base = base.where(AuditLog.admin_user_id == admin_user_id)
    if admin_username:
        admin_ids_subq = (
            select(AdminUser.id)
            .where(AdminUser.username.ilike(f"%{admin_username}%"))
            .scalar_subquery()
        )
        base = base.where(AuditLog.admin_user_id.in_(admin_ids_subq))

    start, end = _parse_dates(start_date, end_date)
    if start:
        base = base.where(AuditLog.created_at >= start)
    if end:
        base = base.where(AuditLog.created_at <= end)

    count_stmt = select(func.count()).select_from(base.subquery())
    total = (await session.execute(count_stmt)).scalar() or 0

    stmt = base.order_by(AuditLog.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size)
    result = await session.execute(stmt)
    logs = result.scalars().all()

    items = [await _build_log_response(session, log) for log in logs]
    return AuditLogListResponse(items=items, total=total, page=page, page_size=page_size)


# ─── Export Audit Logs (Excel) ── MUST be before {log_id} route ──

@router.get("/logs/export")
async def export_audit_logs(
    action: str | None = Query(None),
    module: str | None = Query(None),
    admin_user_id: int | None = Query(None),
    admin_username: str | None = Query(None, description="Admin username (partial match)"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("audit_log.export")),
):
    base = select(AuditLog)

    if action:
        base = base.where(AuditLog.action == action)
    if module:
        base = base.where(AuditLog.module == module)
    if admin_user_id:
        base = base.where(AuditLog.admin_user_id == admin_user_id)
    if admin_username:
        admin_ids_subq = (
            select(AdminUser.id)
            .where(AdminUser.username.ilike(f"%{admin_username}%"))
            .scalar_subquery()
        )
        base = base.where(AuditLog.admin_user_id.in_(admin_ids_subq))

    start, end = _parse_dates(start_date, end_date)
    if start:
        base = base.where(AuditLog.created_at >= start)
    if end:
        base = base.where(AuditLog.created_at <= end)

    stmt = base.order_by(AuditLog.created_at.desc()).limit(5000)
    result = await session.execute(stmt)
    logs = result.scalars().all()

    # Build Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    headers = ["ID", "Admin User ID", "Username", "Action", "Module",
               "Resource Type", "Resource ID", "IP Address", "Description", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, log in enumerate(logs, 2):
        admin_user = await session.get(AdminUser, log.admin_user_id) if log.admin_user_id else None
        ws.cell(row=row_idx, column=1, value=log.id)
        ws.cell(row=row_idx, column=2, value=log.admin_user_id)
        ws.cell(row=row_idx, column=3, value=admin_user.username if admin_user else "")
        ws.cell(row=row_idx, column=4, value=log.action)
        ws.cell(row=row_idx, column=5, value=log.module)
        ws.cell(row=row_idx, column=6, value=log.resource_type)
        ws.cell(row=row_idx, column=7, value=log.resource_id)
        ws.cell(row=row_idx, column=8, value=log.ip_address)
        ws.cell(row=row_idx, column=9, value=log.description)
        ws.cell(row=row_idx, column=10, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "")

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().strftime("%Y-%m-%d")
    filename = f"audit_logs_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type=EXCEL_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Get Audit Log Detail ────────────────────────────────────────

@router.get("/logs/{log_id}", response_model=AuditLogResponse)
async def get_audit_log(
    log_id: int,
    session: AsyncSession = Depends(get_session),
    current_user: AdminUser = Depends(PermissionChecker("audit_log.view")),
):
    from fastapi import HTTPException
    log = await session.get(AuditLog, log_id)
    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")
    return await _build_log_response(session, log)
